"""Employee Import/Export — Excel bulk operations"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from typing import Optional
from datetime import datetime, timezone
import uuid
import io

from database import db
from auth import get_current_user, require_admin, log_activity, hash_password, check_department_access

router = APIRouter()


COLUMN_MAP = {
    "الاسم": "name", "Name": "name", "name": "name",
    "المسمى الوظيفي": "job_title", "Job Title": "job_title", "job_title": "job_title",
    "رقم الموظف": "employee_number", "Employee Number": "employee_number", "employee_number": "employee_number",
    "رقم الهوية": "national_id", "National ID": "national_id", "national_id": "national_id",
    "الجوال": "contact_phone", "Phone": "contact_phone", "contact_phone": "contact_phone",
    "الوردية": "shift", "Shift": "shift", "shift": "shift",
    "الموقع": "location", "Location": "location", "location": "location",
    "نوع التوظيف": "employment_type", "Employment Type": "employment_type", "employment_type": "employment_type",
    "نوع العمل": "work_type", "Work Type": "work_type", "work_type": "work_type",
    "أيام الراحة": "weekly_rest", "Rest Days": "weekly_rest", "weekly_rest": "weekly_rest",
    "انتهاء العقد": "contract_end", "Contract End": "contract_end", "contract_end": "contract_end",
    "القسم": "department", "Department": "department", "department": "department",
}

EXPORT_COLUMNS = [
    ("الاسم", "name"),
    ("رقم الموظف", "employee_number"),
    ("رقم الهوية", "national_id"),
    ("المسمى الوظيفي", "job_title"),
    ("الجوال", "contact_phone"),
    ("الوردية", "shift"),
    ("الموقع", "location"),
    ("نوع التوظيف", "employment_type"),
    ("نوع العمل", "work_type"),
    ("أيام الراحة", "weekly_rest"),
    ("انتهاء العقد", "contract_end"),
    ("القسم", "department"),
    ("الحالة", "is_active"),
]


@router.post("/employees/import")
async def import_employees(
    file: UploadFile = File(...),
    department: str = Form(...),
    user: dict = Depends(get_current_user)
):
    """استيراد موظفين من ملف Excel"""
    if not check_department_access(user, department):
        raise HTTPException(status_code=403, detail="لا يمكنك الوصول لهذه الإدارة")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="يرجى رفع ملف Excel (.xlsx)")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة openpyxl غير مثبتة")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="الملف فارغ أو لا يحتوي على بيانات")

    # Map headers
    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    mapped_headers = []
    for h in raw_headers:
        mapped = COLUMN_MAP.get(h)
        if not mapped:
            for key, val in COLUMN_MAP.items():
                if key.lower() == h.lower():
                    mapped = val
                    break
        mapped_headers.append(mapped)

    if "name" not in mapped_headers:
        raise HTTPException(status_code=400, detail="الملف لا يحتوي على عمود 'الاسم' أو 'Name'")

    created = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            data = {}
            for col_idx, val in enumerate(row):
                if col_idx < len(mapped_headers) and mapped_headers[col_idx]:
                    field = mapped_headers[col_idx]
                    data[field] = str(val).strip() if val is not None else ""

            name = data.get("name", "").strip()
            if not name:
                skipped += 1
                continue

            # Check duplicate by national_id or employee_number
            nid = data.get("national_id", "").strip()
            emp_num = data.get("employee_number", "").strip()
            if nid:
                existing = await db.employees.find_one({"national_id": nid, "department": department})
                if existing:
                    skipped += 1
                    continue
            if emp_num:
                existing = await db.employees.find_one({"employee_number": emp_num, "department": department})
                if existing:
                    skipped += 1
                    continue

            # Parse rest days
            rest_str = data.get("weekly_rest", "")
            rest_days = [d.strip() for d in rest_str.replace("،", ",").split(",") if d.strip()] if rest_str else []

            emp_id = str(uuid.uuid4())
            if not emp_num:
                emp_num = str(1000 + created + row_idx)

            emp_doc = {
                "id": emp_id,
                "name": name,
                "job_title": data.get("job_title", "موظف"),
                "department": department,
                "employee_number": emp_num,
                "national_id": nid or None,
                "contact_phone": data.get("contact_phone", ""),
                "shift": data.get("shift", ""),
                "location": data.get("location", ""),
                "employment_type": data.get("employment_type", "permanent"),
                "work_type": data.get("work_type", "field"),
                "weekly_rest": rest_str,
                "rest_days": rest_days,
                "contract_end": data.get("contract_end", "") or None,
                "work_tasks": "",
                "is_tasked": False,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.employees.insert_one(emp_doc)

            # Auto-create user account if national_id exists
            if nid:
                existing_user = await db.users.find_one({"national_id": nid})
                if not existing_user:
                    user_doc = {
                        "id": str(uuid.uuid4()),
                        "email": None,
                        "national_id": nid,
                        "password": hash_password(emp_num),
                        "name": name,
                        "role": "field_staff",
                        "department": department,
                        "account_status": "pending",
                        "must_change_pin": True,
                        "failed_attempts": 0,
                        "employee_id": emp_id,
                        "is_active": False,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    }
                    await db.users.insert_one(user_doc)

            created += 1
        except Exception as e:
            errors.append(f"سطر {row_idx}: {str(e)[:80]}")

    await log_activity("استيراد موظفين", user, department, f"تم استيراد {created} موظف، تخطي {skipped}")

    return {
        "message": f"تم استيراد {created} موظف بنجاح",
        "created": created,
        "skipped": skipped,
        "errors": errors[:10],
        "total_rows": len(rows) - 1,
    }


@router.get("/employees/export")
async def export_employees(department: Optional[str] = None, user: dict = Depends(get_current_user)):
    """تصدير موظفين إلى Excel"""
    import openpyxl
    from fastapi.responses import StreamingResponse

    query = {}
    if department:
        if not check_department_access(user, department):
            raise HTTPException(status_code=403, detail="لا يمكنك الوصول لهذه الإدارة")
        query["department"] = department

    employees = await db.employees.find(query, {"_id": 0}).to_list(5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الموظفين"
    ws.sheet_view.rightToLeft = True

    # Headers
    headers = [col[0] for col in EXPORT_COLUMNS]
    ws.append(headers)

    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="004D38", end_color="004D38", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 18

    # Data
    for emp in employees:
        row = []
        for _, field in EXPORT_COLUMNS:
            val = emp.get(field, "")
            if field == "is_active":
                val = "نشط" if val else "غير نشط"
            if field == "employment_type":
                val = {"permanent": "دائم", "seasonal": "موسمي", "temporary": "مؤقت"}.get(val, val)
            if field == "work_type":
                val = {"field": "ميداني", "office": "إداري"}.get(val, val)
            row.append(str(val) if val else "")
        ws.append(row)

    # Style data rows
    data_align = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(EXPORT_COLUMNS)):
        for cell in row:
            cell.alignment = data_align
            cell.border = thin_border

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    dept_name = department or "all"
    filename = f"employees_{dept_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/employees/export/template")
async def download_import_template():
    """تحميل قالب Excel للاستيراد"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قالب الموظفين"
    ws.sheet_view.rightToLeft = True

    headers = ["الاسم", "رقم الموظف", "رقم الهوية", "المسمى الوظيفي", "الجوال", "الوردية", "الموقع", "نوع التوظيف", "نوع العمل", "أيام الراحة", "انتهاء العقد"]

    ws.append(headers)

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="004D38", end_color="004D38", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 20

    # Sample row
    ws.append(["أحمد محمد", "1001", "1234567890", "مراقب بوابة", "0512345678", "الأولى", "بوابة 15", "دائم", "ميداني", "الجمعة، السبت", "2026-12-31"])

    sample_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[2]:
        cell.alignment = sample_align
        cell.border = thin_border
        cell.font = Font(color="999999", italic=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="employee_import_template.xlsx"'}
    )
