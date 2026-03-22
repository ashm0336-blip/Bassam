"""Daily Statistics — CRUD + Import/Export for mosque crowd data"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from typing import Optional
from datetime import datetime, timezone
import uuid
import io
import logging

from database import db
from auth import get_current_user, log_activity

router = APIRouter()
logger = logging.getLogger(__name__)

# ─── Field definitions ───────────────────────────────────────────
HARAM_FIELDS = [
    ("haram_worshippers", "المصلين", "Worshippers"),
    ("haram_umrah", "المعتمرين", "Umrah Pilgrims"),
    ("haram_hijr_ismail", "حجر إسماعيل", "Hijr Ismail"),
    ("haram_carts", "العربات", "Carts"),
]

NABAWI_FIELDS = [
    ("nabawi_worshippers", "المصلين", "Worshippers"),
    ("nabawi_rawdah_men_published", "الروضة رجال - منشور", "Rawdah Men Published"),
    ("nabawi_rawdah_men_reserved", "الروضة رجال - محجوز", "Rawdah Men Reserved"),
    ("nabawi_rawdah_men_actual", "الروضة رجال - فعلي", "Rawdah Men Actual"),
    ("nabawi_rawdah_women_published", "الروضة نساء - منشور", "Rawdah Women Published"),
    ("nabawi_rawdah_women_reserved", "الروضة نساء - محجوز", "Rawdah Women Reserved"),
    ("nabawi_rawdah_women_actual", "الروضة نساء - فعلي", "Rawdah Women Actual"),
    ("nabawi_salam_corridor", "ممر السلام", "Salam Corridor"),
]

ALL_NUMERIC_FIELDS = [f[0] for f in HARAM_FIELDS + NABAWI_FIELDS]

# Column map for Excel import
IMPORT_COLUMN_MAP = {
    # Arabic headers
    "التاريخ": "_date_hijri",
    # Haram
    "المصلين": "haram_worshippers",
    "المعتمرين": "haram_umrah",
    "حجر إسماعيل": "haram_hijr_ismail",
    "العربات": "haram_carts",
    # Nabawi
    "ممر السلام": "nabawi_salam_corridor",
}

# These need context-aware mapping (under which parent header)
NABAWI_RAWDAH_HEADERS = {
    "منشور": "published",
    "محجوز": "reserved",
    "فعلي": "actual",
}


def _parse_number(val) -> Optional[float]:
    """Parse a numeric value from Excel, handling Arabic/comma formats."""
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("٬", "")
    if not s or s == "-" or s == "None":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _safe_doc(doc: dict) -> dict:
    """Remove _id and ensure all numeric fields are present."""
    doc.pop("_id", None)
    for f in ALL_NUMERIC_FIELDS:
        if f not in doc:
            doc[f] = None
    return doc


# ─── CRUD Endpoints ──────────────────────────────────────────────

@router.get("/daily-stats")
async def list_daily_stats(
    month: Optional[str] = None,
    year: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    mosque: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    user: dict = Depends(get_current_user),
):
    """List daily stats with filters. month/year are Hijri."""
    query = {}

    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        query["date_hijri"] = date_q
    elif month and year:
        prefix = f"{year}-{month.zfill(2)}"
        query["date_hijri"] = {"$regex": f"^{prefix}"}
    elif year:
        query["date_hijri"] = {"$regex": f"^{year}"}

    total = await db.daily_stats.count_documents(query)
    skip = (page - 1) * limit

    cursor = db.daily_stats.find(query, {"_id": 0}).sort("date_hijri", 1).skip(skip).limit(limit)
    items = await cursor.to_list(limit)

    # If mosque filter, zero out irrelevant fields in response
    if mosque == "haram":
        for item in items:
            for f, _, _ in NABAWI_FIELDS:
                item[f] = None
    elif mosque == "nabawi":
        for item in items:
            for f, _, _ in HARAM_FIELDS:
                item[f] = None

    return {
        "items": [_safe_doc(i) for i in items],
        "total": total,
        "page": page,
        "pages": max(1, (total + limit - 1) // limit),
    }


@router.get("/daily-stats/summary")
async def get_stats_summary(
    month: Optional[str] = None,
    year: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Get summary stats for a period including min/max with dates."""
    match = {}
    if month and year:
        prefix = f"{year}-{month.zfill(2)}"
        match["date_hijri"] = {"$regex": f"^{prefix}"}
    elif year:
        match["date_hijri"] = {"$regex": f"^{year}"}

    pipeline = [{"$match": match}] if match else []

    # Build aggregation for all numeric fields
    group_stage = {"_id": None, "count": {"$sum": 1}}
    for f in ALL_NUMERIC_FIELDS:
        group_stage[f"sum_{f}"] = {"$sum": {"$ifNull": [f"${f}", 0]}}
        group_stage[f"max_{f}"] = {"$max": f"${f}"}
        group_stage[f"min_{f}"] = {"$min": f"${f}"}
        group_stage[f"avg_{f}"] = {"$avg": f"${f}"}

    pipeline.append({"$group": group_stage})
    pipeline.append({"$project": {"_id": 0}})

    result = await db.daily_stats.aggregate(pipeline).to_list(1)
    if not result:
        return {"count": 0}

    summary = result[0]
    # Round averages
    for key in list(summary.keys()):
        if key.startswith("avg_") and summary[key] is not None:
            summary[key] = round(summary[key], 1)

    # Find dates for max/min of key fields
    key_fields = ["haram_worshippers", "nabawi_worshippers"]
    query = match if match else {}
    for kf in key_fields:
        max_val = summary.get(f"max_{kf}")
        min_val = summary.get(f"min_{kf}")
        if max_val is not None:
            doc = await db.daily_stats.find_one({**query, kf: max_val}, {"_id": 0, "date_hijri": 1})
            summary[f"max_{kf}_date"] = doc["date_hijri"] if doc else None
        if min_val is not None and min_val > 0:
            doc = await db.daily_stats.find_one({**query, kf: min_val}, {"_id": 0, "date_hijri": 1})
            summary[f"min_{kf}_date"] = doc["date_hijri"] if doc else None

    return summary


@router.get("/daily-stats/{stat_id}")
async def get_daily_stat(stat_id: str, user: dict = Depends(get_current_user)):
    doc = await db.daily_stats.find_one({"id": stat_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    return _safe_doc(doc)


@router.post("/daily-stats")
async def create_daily_stat(data: dict, user: dict = Depends(get_current_user)):
    """Create or update daily stat for a date."""
    date_hijri = data.get("date_hijri", "").strip().replace("/", "-")
    date_gregorian = data.get("date_gregorian", "").strip()

    if not date_hijri:
        raise HTTPException(status_code=400, detail="التاريخ الهجري مطلوب")

    # Normalize: 1446-9-1 → 1446-09-01
    parts = date_hijri.split("-")
    if len(parts) == 3:
        date_hijri = f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"

    # Check if record exists for this date
    existing = await db.daily_stats.find_one({"date_hijri": date_hijri}, {"_id": 0})

    numeric_data = {}
    for f in ALL_NUMERIC_FIELDS:
        val = data.get(f)
        if val is not None and val != "" and val != "null":
            numeric_data[f] = _parse_number(val)

    if existing:
        # Update existing - merge fields
        update_set = {"updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": user["id"]}
        if date_gregorian:
            update_set["date_gregorian"] = date_gregorian
        for f, v in numeric_data.items():
            if v is not None:
                update_set[f] = v

        await db.daily_stats.update_one(
            {"date_hijri": date_hijri},
            {"$set": update_set}
        )
        updated = await db.daily_stats.find_one({"date_hijri": date_hijri}, {"_id": 0})
        await log_activity("تحديث إحصائية يومية", user, date_hijri)
        return _safe_doc(updated)

    # Create new
    doc = {
        "id": str(uuid.uuid4()),
        "date_hijri": date_hijri,
        "date_gregorian": date_gregorian or "",
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    for f in ALL_NUMERIC_FIELDS:
        doc[f] = numeric_data.get(f)

    await db.daily_stats.insert_one({**doc})
    doc.pop("_id", None)
    await log_activity("إنشاء إحصائية يومية", user, date_hijri)
    return _safe_doc(doc)


@router.post("/daily-stats/fix-dates")
async def fix_date_formats(user: dict = Depends(get_current_user)):
    """Fix dates that were stored without zero-padding (e.g., 1446-9-1 → 1446-09-01)."""
    all_docs = await db.daily_stats.find({}, {"_id": 0, "date_hijri": 1}).to_list(10000)
    fixed = 0
    for doc in all_docs:
        old_date = doc.get("date_hijri", "")
        parts = old_date.replace("/", "-").split("-")
        if len(parts) == 3:
            new_date = f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            if new_date != old_date:
                # Check if normalized date already exists
                existing = await db.daily_stats.find_one({"date_hijri": new_date})
                if existing:
                    # Merge into existing record
                    old_doc = await db.daily_stats.find_one({"date_hijri": old_date}, {"_id": 0})
                    merge = {}
                    for f in ALL_NUMERIC_FIELDS:
                        old_val = old_doc.get(f)
                        ex_val = existing.get(f)
                        if old_val is not None and (ex_val is None):
                            merge[f] = old_val
                    if merge:
                        await db.daily_stats.update_one({"date_hijri": new_date}, {"$set": merge})
                    await db.daily_stats.delete_one({"date_hijri": old_date})
                else:
                    await db.daily_stats.update_one({"date_hijri": old_date}, {"$set": {"date_hijri": new_date}})
                fixed += 1
    await log_activity("إصلاح تواريخ الإحصائيات", user, details=f"تم إصلاح {fixed} سجل")
    return {"message": f"تم إصلاح {fixed} تاريخ", "fixed": fixed}



@router.put("/daily-stats/{stat_id}")
async def update_daily_stat(stat_id: str, data: dict, user: dict = Depends(get_current_user)):
    existing = await db.daily_stats.find_one({"id": stat_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="السجل غير موجود")

    update_set = {"updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": user["id"]}

    if "date_hijri" in data:
        update_set["date_hijri"] = data["date_hijri"]
    if "date_gregorian" in data:
        update_set["date_gregorian"] = data["date_gregorian"]

    for f in ALL_NUMERIC_FIELDS:
        if f in data:
            val = _parse_number(data[f])
            update_set[f] = val

    await db.daily_stats.update_one({"id": stat_id}, {"$set": update_set})
    updated = await db.daily_stats.find_one({"id": stat_id}, {"_id": 0})
    await log_activity("تحديث إحصائية يومية", user, stat_id)
    return _safe_doc(updated)


@router.delete("/daily-stats/{stat_id}")
async def delete_daily_stat(stat_id: str, user: dict = Depends(get_current_user)):
    result = await db.daily_stats.delete_one({"id": stat_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    await log_activity("حذف إحصائية يومية", user, stat_id)
    return {"message": "تم الحذف بنجاح"}


# ─── Import / Export ──────────────────────────────────────────────

@router.post("/daily-stats/import")
async def import_daily_stats(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """Import daily stats from Excel file."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="يرجى رفع ملف Excel (.xlsx)")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة openpyxl غير مثبتة")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        raise HTTPException(status_code=400, detail="الملف يحتاج على الأقل 3 صفوف (عناوين + بيانات)")

    # Parse headers - the spreadsheet has merged headers across 2-3 rows
    # Row 1: Parent categories (المسجد الحرام, المسجد النبوي)
    # Row 2: Sub-categories (المصلين, الروضة رجال, etc.)
    # Row 3: Sub-sub-categories (منشور, محجوز, فعلي) - optional
    # We need smart column mapping

    # Strategy: scan first 3 rows to build column mapping
    header_rows = rows[:3] if len(rows) >= 3 else rows[:2]
    num_cols = max(len(r) for r in header_rows) if header_rows else 0

    # Build context per column
    col_mapping = {}
    parent_context = [None] * num_cols  # Track parent header per column
    sub_context = [None] * num_cols

    # Row 0: parent categories
    if len(header_rows) > 0:
        last_parent = None
        for ci, val in enumerate(header_rows[0]):
            if val and str(val).strip():
                last_parent = str(val).strip()
            parent_context[ci] = last_parent

    # Row 1: sub-categories
    if len(header_rows) > 1:
        last_sub = None
        for ci, val in enumerate(header_rows[1]):
            if val and str(val).strip():
                last_sub = str(val).strip()
            sub_context[ci] = last_sub

    # Row 2: sub-sub-categories (منشور/محجوز/فعلي)
    sub_sub = [None] * num_cols
    data_start_row = 2
    if len(header_rows) > 2:
        has_sub_sub = False
        for ci, val in enumerate(header_rows[2]):
            if val and str(val).strip() in NABAWI_RAWDAH_HEADERS:
                has_sub_sub = True
                sub_sub[ci] = str(val).strip()
        if has_sub_sub:
            data_start_row = 3

    # Now map columns to field names
    for ci in range(num_cols):
        parent = parent_context[ci] or ""
        sub = sub_context[ci] or ""
        ss = sub_sub[ci] or ""

        # Date column
        if "التاريخ" in sub or "التاريخ" in parent:
            col_mapping[ci] = "_date"
            continue

        is_haram = "الحرام" in parent
        is_nabawi = "النبوي" in parent

        if is_haram:
            if "المصلين" in sub or "المصلين" in parent:
                if "المصلين" == sub:
                    col_mapping[ci] = "haram_worshippers"
            if "المعتمرين" in sub:
                col_mapping[ci] = "haram_umrah"
            if "حجر" in sub or "إسماعيل" in sub:
                col_mapping[ci] = "haram_hijr_ismail"
            if "العربات" in sub:
                col_mapping[ci] = "haram_carts"
            if "المصلين" in sub and ci not in col_mapping:
                col_mapping[ci] = "haram_worshippers"

        elif is_nabawi:
            if "المصلين" in sub:
                col_mapping[ci] = "nabawi_worshippers"
            elif "ممر" in sub or "السلام" in sub:
                col_mapping[ci] = "nabawi_salam_corridor"
            elif "الروضة" in sub and "رجال" in sub:
                key = NABAWI_RAWDAH_HEADERS.get(ss, "")
                if key == "published":
                    col_mapping[ci] = "nabawi_rawdah_men_published"
                elif key == "reserved":
                    col_mapping[ci] = "nabawi_rawdah_men_reserved"
                elif key == "actual":
                    col_mapping[ci] = "nabawi_rawdah_men_actual"
                elif not ss:
                    col_mapping[ci] = "nabawi_rawdah_men_published"
            elif "الروضة" in sub and "نساء" in sub:
                key = NABAWI_RAWDAH_HEADERS.get(ss, "")
                if key == "published":
                    col_mapping[ci] = "nabawi_rawdah_women_published"
                elif key == "reserved":
                    col_mapping[ci] = "nabawi_rawdah_women_reserved"
                elif key == "actual":
                    col_mapping[ci] = "nabawi_rawdah_women_actual"
                elif not ss:
                    col_mapping[ci] = "nabawi_rawdah_women_published"

    # Fallback: try simple single-row header mapping
    if not col_mapping:
        if len(header_rows) > 0:
            for ci, val in enumerate(header_rows[0]):
                if val and str(val).strip() in IMPORT_COLUMN_MAP:
                    mapped = IMPORT_COLUMN_MAP[str(val).strip()]
                    if mapped == "_date_hijri":
                        col_mapping[ci] = "_date"
                    else:
                        col_mapping[ci] = mapped

    if "_date" not in col_mapping.values():
        # Try to find date column by checking which column has date-like values
        for ci in range(num_cols):
            for row in rows[data_start_row:data_start_row+3]:
                if ci < len(row) and row[ci]:
                    val = str(row[ci]).strip()
                    if "-" in val and len(val) >= 8:
                        col_mapping[ci] = "_date"
                        break
            if "_date" in col_mapping.values():
                break

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(rows[data_start_row:], start=data_start_row + 1):
        try:
            record = {}
            date_val = None

            for ci, val in enumerate(row):
                field = col_mapping.get(ci)
                if not field:
                    continue
                if field == "_date":
                    date_val = str(val).strip() if val else None
                elif field in ALL_NUMERIC_FIELDS:
                    record[field] = _parse_number(val)

            if not date_val:
                skipped += 1
                continue

            # Normalize date format: 1446/9/1 → 1446-09-01
            date_hijri = date_val.replace("/", "-")
            parts = date_hijri.split("-")
            if len(parts) == 3:
                date_hijri = f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"

            # Check existing
            existing = await db.daily_stats.find_one({"date_hijri": date_hijri})

            if existing:
                update_set = {"updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": user["id"]}
                for f, v in record.items():
                    if v is not None:
                        update_set[f] = v
                await db.daily_stats.update_one({"date_hijri": date_hijri}, {"$set": update_set})
                updated += 1
            else:
                doc = {
                    "id": str(uuid.uuid4()),
                    "date_hijri": date_hijri,
                    "date_gregorian": "",
                    "created_by": user["id"],
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
                for f in ALL_NUMERIC_FIELDS:
                    doc[f] = record.get(f)
                await db.daily_stats.insert_one({**doc})
                created += 1

        except Exception as e:
            errors.append(f"صف {row_idx}: {str(e)[:80]}")

    await log_activity("استيراد إحصائيات يومية", user, target=None,
                       details=f"جديد: {created}, محدث: {updated}, تخطي: {skipped}")

    return {
        "message": f"تم استيراد {created} سجل جديد وتحديث {updated} سجل",
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:10],
        "total_rows": len(rows) - data_start_row,
        "columns_mapped": {str(k): v for k, v in col_mapping.items()},
    }


@router.get("/daily-stats/export/data")
async def export_daily_stats(
    month: Optional[str] = None,
    year: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Export daily stats to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    query = {}
    if month and year:
        prefix = f"{year}-{month.zfill(2)}"
        query["date_hijri"] = {"$regex": f"^{prefix}"}
    elif year:
        query["date_hijri"] = {"$regex": f"^{year}"}

    items = await db.daily_stats.find(query, {"_id": 0}).sort("date_hijri", 1).to_list(5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الإحصائيات اليومية"
    ws.sheet_view.rightToLeft = True

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill_haram = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    header_fill_nabawi = PatternFill(start_color="1e8449", end_color="1e8449", fill_type="solid")
    header_fill_date = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # Row 1: Parent headers
    ws.merge_cells('A1:B1')
    ws['A1'] = 'التاريخ'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill_date
    ws['A1'].alignment = center_align

    ws.merge_cells('C1:F1')
    ws['C1'] = 'المسجد الحرام'
    ws['C1'].font = header_font
    ws['C1'].fill = header_fill_haram
    ws['C1'].alignment = center_align

    ws.merge_cells('G1:N1')
    ws['G1'] = 'المسجد النبوي'
    ws['G1'].font = header_font
    ws['G1'].fill = header_fill_nabawi
    ws['G1'].alignment = center_align

    # Row 2: Sub headers
    sub_headers = [
        ("هجري", header_fill_date),
        ("ميلادي", header_fill_date),
        ("المصلين", header_fill_haram),
        ("المعتمرين", header_fill_haram),
        ("حجر إسماعيل", header_fill_haram),
        ("العربات", header_fill_haram),
        ("المصلين", header_fill_nabawi),
        ("الروضة رجال - منشور", header_fill_nabawi),
        ("الروضة رجال - محجوز", header_fill_nabawi),
        ("الروضة رجال - فعلي", header_fill_nabawi),
        ("الروضة نساء - منشور", header_fill_nabawi),
        ("الروضة نساء - محجوز", header_fill_nabawi),
        ("الروضة نساء - فعلي", header_fill_nabawi),
        ("ممر السلام", header_fill_nabawi),
    ]

    for ci, (header, fill) in enumerate(sub_headers, 1):
        cell = ws.cell(row=2, column=ci, value=header)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # Data rows
    field_order = [
        "date_hijri", "date_gregorian",
        "haram_worshippers", "haram_umrah", "haram_hijr_ismail", "haram_carts",
        "nabawi_worshippers",
        "nabawi_rawdah_men_published", "nabawi_rawdah_men_reserved", "nabawi_rawdah_men_actual",
        "nabawi_rawdah_women_published", "nabawi_rawdah_women_reserved", "nabawi_rawdah_women_actual",
        "nabawi_salam_corridor",
    ]

    data_align = Alignment(horizontal="center", vertical="center")
    for item in items:
        row_data = []
        for f in field_order:
            val = item.get(f)
            if val is None:
                row_data.append("")
            elif isinstance(val, (int, float)):
                row_data.append(val)
            else:
                row_data.append(str(val))
        ws.append(row_data)

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=len(field_order)):
        for cell in row:
            cell.alignment = data_align
            cell.border = thin_border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    period = f"{year or 'all'}_{month or 'all'}"
    filename = f"daily_stats_{period}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/daily-stats/export/template")
async def download_stats_template(user: dict = Depends(get_current_user)):
    """Download Excel template for importing daily stats."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قالب الإحصائيات"
    ws.sheet_view.rightToLeft = True

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill_haram = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    header_fill_nabawi = PatternFill(start_color="1e8449", end_color="1e8449", fill_type="solid")
    header_fill_date = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # Row 1: Parent headers
    ws.merge_cells('A1:A2')
    ws['A1'] = 'التاريخ'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill_date
    ws['A1'].alignment = center_align

    ws.merge_cells('B1:E1')
    ws['B1'] = 'المسجد الحرام'
    ws['B1'].font = header_font
    ws['B1'].fill = header_fill_haram
    ws['B1'].alignment = center_align

    ws.merge_cells('F1:M1')
    ws['F1'] = 'المسجد النبوي'
    ws['F1'].font = header_font
    ws['F1'].fill = header_fill_nabawi
    ws['F1'].alignment = center_align

    # Row 2: Sub headers
    sub_headers = [
        (None, None),  # A already merged
        ("المصلين", header_fill_haram),
        ("المعتمرين", header_fill_haram),
        ("حجر إسماعيل", header_fill_haram),
        ("العربات", header_fill_haram),
        ("المصلين", header_fill_nabawi),
        ("الروضة رجال - منشور", header_fill_nabawi),
        ("الروضة رجال - محجوز", header_fill_nabawi),
        ("الروضة رجال - فعلي", header_fill_nabawi),
        ("الروضة نساء - منشور", header_fill_nabawi),
        ("الروضة نساء - محجوز", header_fill_nabawi),
        ("الروضة نساء - فعلي", header_fill_nabawi),
        ("ممر السلام", header_fill_nabawi),
    ]

    for ci, item in enumerate(sub_headers, 1):
        if item[0] is None:
            continue
        cell = ws.cell(row=2, column=ci, value=item[0])
        cell.font = header_font
        cell.fill = item[1]
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = 18

    ws.column_dimensions['A'].width = 14

    # Sample row
    sample = ["1446-09-01", 75000, 40000, 5000, 1200, 60000, 8000, 6000, 7500, 4000, 3000, 3500, 2000]
    for ci, val in enumerate(sample, 1):
        cell = ws.cell(row=3, column=ci, value=val)
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = Font(color="999999", italic=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="daily_stats_template.xlsx"'}
    )
