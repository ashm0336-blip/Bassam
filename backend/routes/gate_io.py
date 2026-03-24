"""Gates Import/Export — Excel bulk operations"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from typing import Optional
from datetime import datetime, timezone
import uuid
import io

from database import db
from auth import get_current_user, require_department_manager, log_activity

router = APIRouter()

GATE_COLUMNS = ["number", "name", "plaza", "gate_type", "direction", "category", "classification", "status", "max_flow"]
GATE_HEADERS_AR = ["رقم الباب", "اسم الباب", "المنطقة", "النوع", "المسار", "الفئة", "التصنيف", "الحالة", "السعة القصوى"]

COLUMN_MAP = {
    "رقم الباب": "number", "Number": "number",
    "اسم الباب": "name", "Name": "name",
    "المنطقة": "plaza", "Plaza": "plaza",
    "النوع": "gate_type", "Type": "gate_type",
    "المسار": "direction", "Direction": "direction",
    "الفئة": "category", "Category": "category",
    "التصنيف": "classification", "Classification": "classification",
    "الحالة": "status", "Status": "status",
    "السعة القصوى": "max_flow", "Max Flow": "max_flow",
}


@router.get("/gates/export")
async def export_gates(user: dict = Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    gates = await db.gates.find({}, {"_id": 0}).sort("number", 1).to_list(5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الأبواب"
    ws.sheet_view.rightToLeft = True

    header_font = Font(name="Cairo", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, header in enumerate(GATE_HEADERS_AR, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, gate in enumerate(gates, 2):
        for col_idx, field in enumerate(GATE_COLUMNS, 1):
            val = gate.get(field, "")
            if isinstance(val, list):
                val = " + ".join(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for col_idx in range(1, len(GATE_COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=gates.xlsx"})


@router.get("/gates/export/template")
async def download_gates_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قالب الأبواب"
    ws.sheet_view.rightToLeft = True

    header_font = Font(name="Cairo", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    example_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, header in enumerate(GATE_HEADERS_AR, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    example = ["1", "باب الملك عبدالعزيز", "الساحة الشمالية", "رئيسي", "دخول", "محرمين", "عام", "مفتوح", "5000"]
    for col_idx, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.fill = example_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for col_idx in range(1, len(GATE_COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=gates_template.xlsx"})


import re
import unicodedata

def _normalize_header(text: str) -> str:
    text = str(text or "").strip()
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r'\s+', ' ', text)
    text = text.replace('\u200f', '').replace('\u200e', '')
    text = text.replace('ة', 'ه').replace('ى', 'ي')
    return text.strip()

COLUMN_MAP_NORMALIZED = {}
for k, v in COLUMN_MAP.items():
    COLUMN_MAP_NORMALIZED[_normalize_header(k)] = v

NAME_KEYWORDS = {"اسم", "الباب", "name", "باب", "اسم الباب", "gate"}

def _find_header_row_and_map(ws):
    for row_idx in range(1, min(ws.max_row + 1, 15)):
        cells = []
        for cell in ws[row_idx]:
            cells.append(str(cell.value or "").strip())
        col_map = {}
        for idx, h in enumerate(cells):
            norm = _normalize_header(h)
            if norm in COLUMN_MAP_NORMALIZED:
                col_map[COLUMN_MAP_NORMALIZED[norm]] = idx
            elif norm in COLUMN_MAP:
                col_map[COLUMN_MAP[norm]] = idx
            elif h in COLUMN_MAP:
                col_map[COLUMN_MAP[h]] = idx
        if "name" in col_map:
            return row_idx, col_map

        for idx, h in enumerate(cells):
            h_lower = h.lower().strip()
            for kw in NAME_KEYWORDS:
                if kw in h_lower or kw in _normalize_header(h):
                    col_map["name"] = idx
                    break
            if "name" in col_map:
                break
        if "name" in col_map:
            for idx2, h2 in enumerate(cells):
                norm2 = _normalize_header(h2)
                if norm2 in COLUMN_MAP_NORMALIZED and COLUMN_MAP_NORMALIZED[norm2] != "name":
                    col_map[COLUMN_MAP_NORMALIZED[norm2]] = idx2
            return row_idx, col_map
    return None, {}


@router.post("/gates/import")
async def import_gates(
    file: UploadFile = File(...),
    mode: str = "skip",
    user: dict = Depends(require_department_manager),
):
    import openpyxl
    import logging
    logger = logging.getLogger("gate_import")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    header_row = None
    col_map = {}
    chosen_ws = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        h_row, c_map = _find_header_row_and_map(ws)
        if h_row and "name" in c_map:
            header_row = h_row
            col_map = c_map
            chosen_ws = ws
            logger.info(f"Found header in sheet '{sheet_name}' row {h_row}, columns: {c_map}")
            break

    if not chosen_ws or "name" not in col_map:
        all_sheets = ", ".join(wb.sheetnames)
        raise HTTPException(
            status_code=400,
            detail=f"لم يتم العثور على عمود 'اسم الباب' في أي شيت. الشيتات: {all_sheets}"
        )

    def _raw(vals, field):
        idx = col_map.get(field, -1)
        if idx == -1 or idx >= len(vals):
            return ""
        return vals[idx]

    all_rows = []
    empty_name_count = 0
    empty_rows = 0
    total_data_rows = 0

    for row in chosen_ws.iter_rows(min_row=header_row + 1, values_only=False):
        vals = []
        for cell in row:
            v = cell.value
            if v is None:
                vals.append("")
            else:
                vals.append(str(v).strip())
        if not any(vals):
            empty_rows += 1
            continue
        total_data_rows += 1
        name = _raw(vals, "name").strip()
        number_str = _raw(vals, "number").strip()
        if not name:
            if number_str:
                name = f"باب {number_str}"
            else:
                empty_name_count += 1
                continue
        raw_max = _raw(vals, "max_flow")
        all_rows.append({
            "number": _raw(vals, "number"),
            "name": name,
            "plaza": _raw(vals, "plaza"),
            "gate_type": _raw(vals, "gate_type"),
            "direction": _raw(vals, "direction"),
            "category": [c.strip() for c in _raw(vals, "category").split("+") if c.strip()] if "category" in col_map and _raw(vals, "category") else [],
            "classification": _raw(vals, "classification"),
            "status": _raw(vals, "status"),
            "max_flow": int(raw_max) if raw_max.isdigit() else 0,
        })

    logger.info(f"Import parse: total_data_rows={total_data_rows}, valid={len(all_rows)}, empty_name={empty_name_count}, empty_rows={empty_rows}")

    if not all_rows:
        raise HTTPException(
            status_code=400,
            detail=f"الملف لا يحتوي على بيانات صالحة. صفوف بيانات: {total_data_rows}، بدون اسم: {empty_name_count}"
        )

    added = 0
    updated = 0
    skipped = 0

    if mode == "replace":
        docs = []
        for gate_data in all_rows:
            gate_data.update({
                "id": str(uuid.uuid4()),
                "current_flow": 0,
                "current_indicator": "",
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            docs.append(gate_data)
        await db.gates.delete_many({})
        await db.gates.insert_many(docs)
        added = len(docs)
    else:
        for gate_data in all_rows:
            name = gate_data["name"]
            number_val = gate_data.get("number", "")
            query = {"name": name}
            if number_val:
                query = {"$or": [{"name": name}, {"number": number_val}]}
            existing = await db.gates.find_one(query, {"_id": 1})
            if existing:
                if mode == "update":
                    await db.gates.update_one({"_id": existing["_id"]}, {"$set": gate_data})
                    updated += 1
                else:
                    skipped += 1
            else:
                gate_data.update({
                    "id": str(uuid.uuid4()),
                    "current_flow": 0,
                    "current_indicator": "",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                })
                await db.gates.insert_one(gate_data)
                added += 1

    parts = []
    if added:
        parts.append(f"إضافة {added}")
    if updated:
        parts.append(f"تحديث {updated}")
    if skipped:
        parts.append(f"تخطي {skipped}")
    summary = "، ".join(parts) if parts else "لا توجد تغييرات"

    await log_activity("استيراد أبواب", user, f"{added+updated}", f"{summary}")
    return {
        "message": f"تم: {summary}",
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "total_rows_in_file": total_data_rows,
        "empty_name_rows": empty_name_count,
        "sheet_used": chosen_ws.title,
    }
